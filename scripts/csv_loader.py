#!/usr/bin/env python3
"""CSV/XLSX loader — read rv-optkb-tool CSV output and group by Idea.

rv-optkb-tool's csv-review/main.py produces CSV files (or xlsx workbooks)
with columns:
    Idea, Thought, Commit URL, Correct?, Why

The "Idea" column is the optimization category → pattern group.
The sheet name (in xlsx) or a project parameter is the project name.

Usage (xlsx):
    groups = load_xlsx("optimization_ideas.xlsx")
    # Returns [GroupInfo, ...], one per sheet

Usage (csv):
    groups = load_csv("openjdk_optimizations.csv", project_name="OpenJDK")
    # Returns [GroupInfo]
"""

import csv
import os
import re
from collections import defaultdict
from typing import Any

# ── Data Structures ───────────────────────────────────────────────────────────

# GitHub commit URL pattern: https://github.com/{owner}/{repo}/commit/{sha}
COMMIT_URL_RE = re.compile(
    r"https://github\.com/([^/]+)/([^/]+)/commit/([a-f0-9]+)"
)


class CommitRow:
    """One row from the CSV: a commit + its optimization idea/thought."""

    __slots__ = ("commit_url", "owner", "repo", "sha",
                 "idea", "thought", "correct", "why")

    def __init__(self, commit_url: str, idea: str, thought: str,
                 correct: str = "", why: str = ""):
        self.commit_url = commit_url
        self.idea = idea
        self.thought = thought
        self.correct = correct
        self.why = why
        # Parse URL
        m = COMMIT_URL_RE.match(commit_url)
        if m:
            self.owner = m.group(1)
            self.repo = m.group(2)
            self.sha = m.group(3)
        else:
            self.owner = ""
            self.repo = ""
            self.sha = ""


class GroupInfo:
    """One Idea group: all commits sharing the same Idea label."""

    def __init__(self, idea: str, project: str, rows: list[CommitRow]):
        self.idea = idea
        self.project = project
        self.rows = rows

    @property
    def count(self) -> int:
        return len(self.rows)

    @property
    def commit_urls(self) -> list[str]:
        return [r.commit_url for r in self.rows]

    def __repr__(self) -> str:
        return (f"GroupInfo(idea={self.idea!r}, project={self.project!r}, "
                f"count={self.count})")


# ── XLSX Reader ───────────────────────────────────────────────────────────────


def load_xlsx(path: str) -> list[GroupInfo]:
    """Read a multi-sheet xlsx workbook.

    Each sheet is one project (sheet name = project name).
    Returns flat list of GroupInfo across all sheets.
    """
    import openpyxl  # lazy import

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    all_groups: list[GroupInfo] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        project = sheet_name.strip()
        # Find header row (look for "Idea" column)
        rows = _xlsx_rows(ws)
        groups = _group_rows(rows, project)
        all_groups.extend(groups)

    wb.close()
    return all_groups


def _xlsx_rows(ws) -> list[dict[str, str]]:
    """Convert worksheet rows to list of dicts, skipping header."""
    # Iterate to find the first row containing "Idea" as a header
    header_idx: dict[str, int] | None = None
    records: list[dict[str, str]] = []

    for row in ws.iter_rows(values_only=True):
        if not row or all(v is None for v in row):
            continue

        # Normalize: None → ""
        row_clean = [str(v).strip() if v is not None else "" for v in row]

        # Detect header: first row where any cell starts with "Idea"
        if header_idx is None:
            for i, cell in enumerate(row_clean):
                if cell.lower() in ("idea", "idea "):
                    header_idx = {
                        "idea": i,
                        "thought": _find_column(row_clean, "thought"),
                        "commit_url": _find_column(row_clean, "commit url"),
                        "correct": _find_column(row_clean, "correct?"),
                        "why": _find_column(row_clean, "why"),
                    }
            continue

        # Data rows: skip empty rows and title rows
        idea = row_clean[header_idx["idea"]] if header_idx["idea"] < len(row_clean) else ""
        if not idea:
            continue  # skip title rows, empty rows

        commit_url = row_clean[header_idx["commit_url"]] if header_idx["commit_url"] < len(row_clean) else ""
        thought = row_clean[header_idx["thought"]] if header_idx["thought"] < len(row_clean) else ""
        correct = row_clean[header_idx["correct"]] if header_idx["correct"] < len(row_clean) else ""
        why = row_clean[header_idx["why"]] if header_idx["why"] < len(row_clean) else ""

        records.append({
            "idea": idea,
            "thought": thought,
            "commit_url": commit_url,
            "correct": correct,
            "why": why,
        })

    return records


def _find_column(row: list[str], name: str) -> int:
    """Find column index by name (case-insensitive, prefix match)."""
    for i, cell in enumerate(row):
        if cell.lower().startswith(name.lower()):
            return i
    return -1


# ── CSV Reader ────────────────────────────────────────────────────────────────


def load_csv(path: str, project_name: str = "") -> list[GroupInfo]:
    """Read a single CSV file.

    Args:
        path: Path to CSV file (utf-8 or utf-8-sig).
        project_name: Override project name. If empty, derived from filename.

    Returns flat list of GroupInfo.
    """
    project = project_name or _project_from_filename(path)

    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)

        # Normalize column names (strip whitespace, lowercase)
        raw_fieldnames = [h.strip() for h in (reader.fieldnames or [])]

        # Build column mapping
        col_map = _build_column_map(raw_fieldnames)

        records: list[dict[str, str]] = []
        for row in reader:
            records.append({
                "idea": row.get(col_map.get("idea", ""), "").strip(),
                "thought": row.get(col_map.get("thought", ""), "").strip(),
                "commit_url": row.get(col_map.get("commit_url", ""), "").strip(),
                "correct": row.get(col_map.get("correct", ""), "").strip(),
                "why": row.get(col_map.get("why", ""), "").strip(),
            })

    return _group_rows(records, project)


def _build_column_map(fieldnames: list[str]) -> dict[str, str]:
    """Map canonical names to actual column names (fuzzy match)."""
    map: dict[str, str] = {}
    for fn in fieldnames:
        fn_lower = fn.lower().strip()
        if fn_lower in ("idea", "idea "):
            map["idea"] = fn
        elif fn_lower in ("thought", "thought "):
            map["thought"] = fn
        elif fn_lower in ("commit url", "commiturl", "commit_url", "commit-url", "commit_url "):
            map["commit_url"] = fn
        elif fn_lower in ("correct", "correct?", "right", "right?"):
            map["correct"] = fn
        elif fn_lower in ("why", "why "):
            map["why"] = fn
    return map


def _project_from_filename(path: str) -> str:
    """Guess project name from CSV filename: 'openjdk_ideas.csv' → 'OpenJDK'."""
    base = os.path.splitext(os.path.basename(path))[0]
    # Strip common suffixes
    for suffix in ["_ideas", "-ideas", "_optimizations", "-optimizations",
                   "_review", "-review", "_classified", "-classified"]:
        if base.lower().endswith(suffix):
            base = base[:-len(suffix)]
            break
    return base


# ── Grouping Logic ────────────────────────────────────────────────────────────


def _group_rows(records: list[dict[str, str]],
                project: str) -> list[GroupInfo]:
    """Group rows by Idea column, return list of GroupInfo.

    Filters out rows marked as incorrect (correct=no).
    Groups with < 2 rows are still returned (caller applies threshold).
    """
    # If records dicts have 'correct' key, filter out "no" answers
    filtered: list[dict[str, str]] = []
    for r in records:
        correct_val = r.get("correct", "").strip().lower()
        if correct_val in ("no", "n", "false", "wrong"):
            continue  # human annotator marked this as incorrect
        filtered.append(r)

    groups: dict[str, list[CommitRow]] = defaultdict(list)
    for r in filtered:
        idea = r["idea"].strip()
        if not idea:
            continue
        commit_url = r["commit_url"].strip()
        if not commit_url:
            continue
        cr = CommitRow(
            commit_url=commit_url,
            idea=idea,
            thought=r["thought"],
            correct=r.get("correct", ""),
            why=r.get("why", ""),
        )
        groups[idea].append(cr)

    result = [GroupInfo(idea=idea, project=project, rows=rows)
              for idea, rows in groups.items()]
    result.sort(key=lambda g: g.count, reverse=True)
    return result


# ── Convenience Loader ────────────────────────────────────────────────────────


def load_any(path: str, project_name: str = "") -> list[GroupInfo]:
    """Auto-detect xlsx vs csv and load."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        return load_xlsx(path)
    elif ext == ".csv":
        return load_csv(path, project_name)
    else:
        raise ValueError(f"Unsupported format: {ext} (use .csv or .xlsx)")


def groups_to_commit_dicts(groups: list[GroupInfo]) -> list[dict[str, Any]]:
    """Convert GroupInfo list to commit dicts compatible with compute_group_stats.

    Returns list of { "url", "idea", "project", "thought", "owner", "repo", "sha" }
    """
    commits = []
    for group in groups:
        for row in group.rows:
            commits.append({
                "url": row.commit_url,
                "idea": row.idea,
                "project": group.project,
                "thought": row.thought,
                "owner": row.owner,
                "repo": row.repo,
                "sha": row.sha,
            })
    return commits
